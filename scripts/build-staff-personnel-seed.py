#!/usr/bin/env python3
"""Build staff-personnel-seed.json and staff-personnel-seed.js from 代號戰力維護 Excel."""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print('pip install openpyxl', file=sys.stderr)
    sys.exit(1)

YEAR = 2026
BASELINE = f'{YEAR}-01-01'
IMPORT_BASELINE_FOREVER = ''
ARROW = r'(?:→|->|➔)'
ROOT = Path(__file__).resolve().parents[1]


def mmdd_to_iso(mmdd: str) -> str:
    mmdd = str(mmdd).zfill(4)
    return f'{YEAR}-{mmdd[:2]}-{mmdd[2:4]}'


def norm_status(s) -> str:
    return str(s or '').strip() or '在職'


def clean_group(g) -> str:
    g = str(g or '').strip()
    if not g or g.lower() == 'none':
        return ''
    return g[:-1] if g.endswith('組') else g


def parse_changes(text):
    if not text or not str(text).strip():
        return []
    entries = []
    for part in re.split(r'[;；]', str(text)):
        part = part.strip()
        if not part:
            continue
        m = re.match(r'^(\d{4})\s*:\s*(.+)$', part)
        if not m:
            continue
        iso = mmdd_to_iso(m.group(1))
        changes = []
        for item in re.split(r'、', m.group(2)):
            item = item.strip()
            if not item:
                continue
            for field, pat in [
                ('level', rf'^戰力\s*(.+?)\s*{ARROW}\s*(.+)$'),
                ('group', rf'^組別\s*(.+?)\s*{ARROW}\s*(.+)$'),
                ('code', rf'^代號\s*(.+?)\s*{ARROW}\s*(.+)$'),
                ('status', rf'^狀態\s*(.+?)\s*{ARROW}\s*(.+)$'),
            ]:
                m2 = re.match(pat, item)
                if m2:
                    changes.append({'field': field, 'from': m2.group(1).strip(), 'to': m2.group(2).strip()})
                    break
        if changes:
            entries.append({'effectiveDate': iso, 'changes': changes})
    entries.sort(key=lambda x: x['effectiveDate'])
    return entries


def build_person(row):
    code, name, group, pid, level, status, hist = (list(row) + [None] * 7)[:7]
    pid = str(pid or '').strip()
    if not pid:
        return None
    final = {
        'code': str(code or '').strip(),
        'fullName': str(name or '').strip(),
        'group': clean_group(group),
        'level': str(level or '').strip(),
        'status': norm_status(status),
        'restrictions': '',
    }
    entries = parse_changes(hist)
    state = dict(final)
    for ent in reversed(entries):
        for ch in reversed(ent['changes']):
            f, val = ch['field'], ch['from']
            if f == 'group':
                state['group'] = clean_group(val)
            elif f == 'level':
                state['level'] = str(val or '').strip()
            elif f == 'status':
                state['status'] = norm_status(val)
            else:
                state[f] = val

    events = []
    first_date = entries[0]['effectiveDate'] if entries else BASELINE
    baseline_date = BASELINE if first_date > BASELINE else first_date
    snap = {
        **state,
        'effectiveDate': IMPORT_BASELINE_FOREVER,
        'fullName': final['fullName'],
        'isImportBaseline': True,
    }
    events.append(snap)
    cur = dict(snap)
    for ent in entries:
        for ch in ent['changes']:
            if ch['field'] == 'code':
                cur['code'] = ch['to']
            elif ch['field'] == 'group':
                cur['group'] = clean_group(ch['to'])
            elif ch['field'] == 'level':
                cur['level'] = str(ch['to'] or '').strip()
            elif ch['field'] == 'status':
                cur['status'] = norm_status(ch['to'])
        cur['effectiveDate'] = ent['effectiveDate']
        cur['fullName'] = final['fullName']
        events.append({**cur})
    deduped = []
    for ev in events:
        if deduped:
            prev = deduped[-1]
            if all(prev.get(k) == ev.get(k) for k in ('code', 'group', 'level', 'status', 'restrictions')) and prev.get('fullName') == ev.get('fullName'):
                continue
        deduped.append(ev)
    return {'personnelId': pid, 'events': deduped}


def main():
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads' / '代號戰力維護_整合版.xlsx'
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet = '整合最新狀態' if '整合最新狀態' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    people = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        p = build_person(row)
        if p:
            people.append(p)
    wb.close()
    for p in people:
        evs = p.get('events') or []
        if evs:
            p['events'] = [{
                **evs[-1],
                'effectiveDate': IMPORT_BASELINE_FOREVER,
                'isImportBaseline': True,
            }]
    payload = {
        'version': 1,
        'source': src.name,
        'historyYear': YEAR,
        'importedAt': datetime.now().isoformat(timespec='seconds'),
        'historyStripped': True,
        'historyBaselineNormalized': True,
        'people': people,
    }
    json_path = ROOT / 'staff-personnel-seed.json'
    js_path = ROOT / 'staff-personnel-seed.js'
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    js_path.write_text(
        '/* Auto-generated — scripts/build-staff-personnel-seed.py */\n'
        f'window.STAFF_PERSONNEL_SEED = {json.dumps(payload, ensure_ascii=False, separators=(",", ":"))};\n',
        encoding='utf-8',
    )
    print(f'Wrote {len(people)} people to {json_path} and {js_path}')


if __name__ == '__main__':
    main()
